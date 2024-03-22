from config import EXCEL_FILEPATH
import unittest
from openpyxl import load_workbook

class TestExcelFile(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.workbook = load_workbook(filename=EXCEL_FILEPATH)
        cls.sheet = cls.workbook.active

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()

    def test_row_count(self):
        """Тест количества строк в итоговой таблице."""
        expected_row_count = 22  
        actual_row_count = self.sheet.max_row
        self.assertEqual(actual_row_count, expected_row_count)

    def test_column_count(self):
        """Тест количества столбцов в итоговой таблице."""
        expected_column_count = 7  
        actual_column_count = self.sheet.max_column
        self.assertEqual(actual_column_count, expected_column_count)

    def test_header_row(self):
        """Проверка на соответствие типов."""
        expected_headers = ['Дата USD/RUB', 'Курс USD/RUB', 'Время USD/RUB', 'Дата JPY/RUB', 'Курс JPY/RUB', 'Время JPY/RUB', 'Результат']
        actual_headers = [self.sheet.cell(row=1, column=col).value for col in range(1, self.sheet.max_column + 1)]
        self.assertEqual(actual_headers, expected_headers)

    def test_second_row_contents_type(self):
        # Тест на содержимое второй ячейки каждого столбца
        types = [str, float, str, str, float, str, float]
        row_data = [self.sheet.cell(row=2, column=col).value for col in range(1, self.sheet.max_column + 1)]

        for i in range(len(row_data)):
            type_match = type(row_data[i]) == types[i]
            assert type_match, "Некорректные данные в таблице!"


if __name__ == '__main__':
    unittest.main()